import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from reports import METRICS
from db import get_plan

# ── Цвета ──────────────────────────────────────────────────────────────────
BLUE_HDR   = "1F4E79"   # тёмно-синий заголовок
BLUE_LIGHT = "BDD7EE"   # светло-синий строки
WHITE      = "FFFFFF"
GRAY_ROW   = "F2F2F2"
GREEN      = "E2EFDA"
ORANGE     = "FCE4D6"

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_cell(ws, row, col, value, bg=BLUE_HDR, font_color=WHITE, bold=True, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=font_color, name="Calibri", size=11)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _border()
    return c

def _data_cell(ws, row, col, value, bg=WHITE, bold=False, align="center", number=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, name="Calibri", size=11)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _border()
    return c


def generate_regular_excel(
    data: dict,
    manager_map: dict,    # {bitrix_id: short_name}
    date_from: str,
    date_to: str,
    proj_label: str,
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Регулярный менеджмент"
    ws.sheet_view.showGridLines = False

    mgr_ids = sorted([mid for mid in data if mid != "total"])

    # ── Строка 1: заголовок отчёта ──────────────────────────────────────
    title = f"Регулярный менеджмент | {proj_label} | {date_from} — {date_to}"
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=2 + len(mgr_ids))
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.fill = PatternFill("solid", fgColor=BLUE_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Строка 2: шапка таблицы ─────────────────────────────────────────
    _hdr_cell(ws, 2, 1, "Метрика", align="left")
    for i, mid in enumerate(mgr_ids):
        _hdr_cell(ws, 2, 2 + i, manager_map.get(mid, str(mid)))
    _hdr_cell(ws, 2, 2 + len(mgr_ids), "Итого", bg="2E75B6")
    ws.row_dimensions[2].height = 22

    # ── Строки данных ────────────────────────────────────────────────────
    total = data.get("total", {})
    for r_idx, (key, label) in enumerate(METRICS):
        row = 3 + r_idx
        bg = GRAY_ROW if r_idx % 2 == 0 else WHITE
        ws.row_dimensions[row].height = 20

        # Метрика
        _data_cell(ws, row, 1, label, bg=bg, align="left")

        # По каждому менеджеру
        for i, mid in enumerate(mgr_ids):
            mdata = data.get(mid, {})
            if key == "cr3":
                dfb = mdata.get("dfb", 0)
                meets = mdata.get("meetings", 0)
                val = f"{round(dfb/meets*100)}%" if meets else "0%"
            else:
                val = mdata.get(key, 0)
            _data_cell(ws, row, 2 + i, val, bg=bg)

        # Итого
        if key == "cr3":
            dfb = total.get("dfb", 0)
            meets = total.get("meetings", 0)
            tot_val = f"{round(dfb/meets*100)}%" if meets else "0%"
            tot_bg = GREEN
        else:
            tot_val = total.get(key, 0)
            tot_bg = GREEN if key == "dfb" else bg
        _data_cell(ws, row, 2 + len(mgr_ids), tot_val,
                   bg=tot_bg, bold=True)

    # ── Ширина столбцов ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    for i in range(len(mgr_ids) + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_planfact_excel(
    data: dict,
    date_from: str,
    date_to: str,
    year: int,
    month: int,
    proj_label: str,
) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "План vs Факт"
    ws.sheet_view.showGridLines = False

    total = data.get("total", {})

    # ── Строка 1: заголовок ──────────────────────────────────────────────
    title = f"План vs Факт | {proj_label} | {date_from} — {date_to}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=13, color=WHITE, name="Calibri")
    c.fill = PatternFill("solid", fgColor=BLUE_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Строка 2: шапка ─────────────────────────────────────────────────
    for col, txt in enumerate(["Метрика", "План", "Факт", "Выполнение"], start=1):
        _hdr_cell(ws, 2, col, txt, align="left" if col == 1 else "center")
    ws.row_dimensions[2].height = 22

    # ── Данные ───────────────────────────────────────────────────────────
    for r_idx, (key, label) in enumerate(METRICS):
        row = 3 + r_idx
        bg = GRAY_ROW if r_idx % 2 == 0 else WHITE
        ws.row_dimensions[row].height = 20

        plan_val = get_plan(year, month, proj_label, key)

        if key == "cr3":
            dfb = total.get("dfb", 0)
            meets = total.get("meetings", 0)
            fact_str = f"{round(dfb/meets*100)}%" if meets else "0%"
            plan_str = f"{int(plan_val)}%" if plan_val else "—"
            pct_str  = ""
        else:
            fact_num = total.get(key, 0)
            fact_str = fact_num
            plan_str = int(plan_val) if plan_val else "—"
            if plan_val and plan_val > 0:
                pct = round(fact_num / plan_val * 100)
                pct_str = f"{pct}%"
                bg_pct = GREEN if pct >= 100 else (ORANGE if pct < 70 else WHITE)
            else:
                pct_str = "—"
                bg_pct  = bg

        _data_cell(ws, row, 1, label, bg=bg, align="left")
        _data_cell(ws, row, 2, plan_str, bg=bg)
        _data_cell(ws, row, 3, fact_str, bg=bg, bold=True)
        _data_cell(ws, row, 4, pct_str,
                   bg=bg_pct if key != "cr3" else bg, bold=True)

    # ── Ширина столбцов ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 24
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
